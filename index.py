import os
import csv 
from openpyxl import load_workbook


# it takes two arguments the list and a tuple of numbers 
# the functions removes each index of the list form the tuple
def list_slicer(state : list, *numbers) :

    initial_state = list(state)

    new_state = state

    for i in numbers: 
        
        # parse the value from the state
        value = state[i]
        # updaitng the new state
        new_state.remove(value)
        # reinitialized the previous state
        state = initial_state

    return new_state

# combining two lists 

def list_combiner(first_list: list, second_list: list) :

    for item in second_list: 
        first_list.append(item)
    return first_list


# execution
def convert_excel_to_csv(filepath):
    # extracting the file name
    filename = os.path.basename(filepath)
    desired_part = filename.split(".")[0]

    # loading the workbook
    wb = load_workbook(filename=filepath)
    sheet = wb.active

    csv_data=[]

    for row in sheet.iter_rows(values_only=True):

        cash_list = []
        # parsing items from the row to convert floats and number to strings
        for item in row : 
            # float handeling
            if type(item) is float:
                
                # this is gonna be the variable that carries our changes make sure that you
                value = int(item)

                # // TODO: check if this column is static 
                if item == row[2] : 

                    value = "0" + str(value)

                # it must be 10 character so we replace the messing chars with zeros
                if item == row[4] or item == row[5]: 

                    value = str(value) 
                    marge_number = 10 - len(value)

                    value = ("0" * marge_number) + value

                cash_list.append(str(value))
            else : 
            # string handling 
                value = str(item)
                cash_list.append(value)

            

        csv_data.append(list(cash_list))

    # this step is important to make the last line of the excel sheet readable
    prefix_list = list(range(13)) # empty list 
    csv_data.append(prefix_list)

    with open (f"{desired_part}.csv", 'w') as csv_obj:
    # the separated char is pip `|`
        writer = csv.writer(csv_obj, delimiter="|")


        count = 0 
        pointer = 0

        for line in csv_data: 

            if (count > 0): 
                count -= 1
                continue

            # prevent conflect with the next line
            copied_line = list(line)

            sliced_line = list_slicer(copied_line, 6, 7, 8, -1)



            for next_line in csv_data[pointer:]: 

                if count == 0: 
                    count += 1
                    continue

                # if the next line doesn't match
                if line[4] != next_line[4] :
                    #write line               

                    pointer += count
                    count -= 1

                    # print(sliced_line)
                    # +0 column written
                    if count == 0: 

                        match sliced_line[-2]: 
                            case "GCB12": 

                                prefix = ["GCB06", "", "GCB03", ""]
                                for item in prefix: 
                                    sliced_line.append(item)

                            case "GCB06": 
                                new_slice = sliced_line[7:]
                                
                                prefix_1 = ["GCB12", ""]
                                prefix_2 = ["GCB03", ""]

                                for item in prefix_1: 
                                    new_slice.append(item)

                                for item in sliced_line[:-2]: 
                                    new_slice.append(item)

                                for item in prefix_2: 
                                    new_slice.append(item)

                                sliced_line = new_slice

                            case "GCB03": 
                                new_slice = sliced_line[7:]
                                
                                prefix = ["GCB12", "", "GCB06", ""]
                            
                                for item in prefix: 
                                    new_slice.append(item)

                                for item in sliced_line[:-2]: 
                                    new_slice.append(item)
                                
                                sliced_line = new_slice

                    # +2 columns written
                    if (count == 1): 

                        match sliced_line[-2]: 

                            case "GCB06": 
                                prefix = [ "GCB03", ""]
                                for item in prefix: 
                                    sliced_line.append(item)

                            case "GCB03": 
                                # one line condition
                                prefix = ["GCB12", ""] if sliced_line[-4] == "GCB06" else ["GCB06", ""]

                                # if the previous was GCB06
                                if sliced_line[-4] == "GCB06": 
                                    new_slice = sliced_line[:7]
                                    
                                    for item in prefix : 
                                        new_slice.append(item) 

                                    for item in sliced_line[-4:]:
                                        new_slice.append(item)
                                    
                                    sliced_line = new_slice

                                else:

                                    new_slice = sliced_line[:9]
                                    
                                    for item in prefix : 
                                        new_slice.append(item) 

                                    for item in sliced_line[-2:]:
                                        new_slice.append(item)
                                    
                                    sliced_line = new_slice
                                
                    # write line

                    writer.writerow(sliced_line)


                    sliced_line = []
                    sliced_next_line = []

                    break 
                

                count += 1
                # append the next_line to the list

                if line[4] == next_line[4]:

                    copied_next_line = list(next_line)

                    sliced_next_line = list_slicer(copied_next_line,0, 1, 2, 3, 4, 5, 8, 6, 7, 9, -1)

                    
                    sliced_line = list_combiner(sliced_line, sliced_next_line)
                    
                




