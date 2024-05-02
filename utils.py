import os
import csv 
from openpyxl import load_workbook
from openpyxl import Workbook
    
# Function to check if all values in a tuple are None
def tuple_all_none(tuple : tuple):
    return all(value is None for value in tuple)

# validate if the row is nonefull "all the row's values are none"
def validaterow(row): 
    non_nullable_indices = [0, 1, 2]

    for index in non_nullable_indices:
        if row[index] == None or row[index] == "None":
            return index
    return True

# converting numbers to letters in an alphabetical way
def convert_to_letters(number):
    if 1 <= number <= 12:
        return chr(ord('a') + number)
    elif number == 0:
        return 'a'
    else:
        return None  
    
    # sort the sheet if it is not on order
def custom_sort(line):
    # Sort by index 5 in ascending order
    value = line[0]
    # Custom order for groups
    group_order = {"GCB12": 0, "GCB06": 1, "GCB03": 2}
    group = line[1]  # Index of the group
    return (value, group_order.get(group, 3))

    # filtring none inside the main array [none, item, none] => [item]
def filter_none(data : list):

    filtered_list = list(filter(lambda x: x != "None" , data))
    return filtered_list

# combining multiples lists
def list_combiner(main_list: list, *prefix: list) :

    formated_list = [item for sublist in prefix for item in sublist]

    for item in formated_list: 
        main_list.append(item)

    return main_list

def merge_lists(list_of_lists):
    merged_lists = {}

    for sublist in list_of_lists:
        key = (sublist[0], sublist[1])
        if key in merged_lists:
            merged_lists[key] += sublist[2]
        else:
            merged_lists[key] = sublist[2]

    result = [[key[0], key[1], value] for key, value in merged_lists.items()]
    return result

    
def convert_excel_to_csv(filepath) : 
    # extracting the file name
    filename = os.path.basename(filepath) 

    desired_part = filename.split(".")[0]

    #create a workbook
    workbook = Workbook()
    worksheet = workbook.active
    raw_data = list()

    wb = load_workbook(filename=filepath) 

     # wb.sheetnames = contains excel files
    worksheet_names = wb.sheetnames

    #  Iterate through each worksheet and combining them
    for worksheet_name in worksheet_names:
        #work sheet
        ws = wb[worksheet_name]

        for row in ws.iter_rows(values_only=True):
        # Append the row to the combined rows list
            raw_data.append(row)

    raw_data = [tup for tup in raw_data if not tuple_all_none(tup)]

    for row_idx, row in enumerate(raw_data, start=1):
        for col_idx, value in enumerate(row, start=1):
            worksheet.cell(row=row_idx, column=col_idx, value=value)

    # Save the workbook to a file
    workbook.save(filename)

    for row in raw_data : 
        # check                 
        validator = validaterow(row) # returns True or None

        if validator is not True: 
            raise ValueError(f"error at line {raw_data.index(row) + 1}  column : {convert_to_letters(validator).upper()}")



    raw_data = [list(t) for t in raw_data]

    # remplacing BTBZER with GCB
    for row in raw_data : 
        row[1] = f'GCB{row[1][-2:]}'
        row[2] = int(str(row[2]).replace(".", ""))

    raw_data = merge_lists(raw_data)
    sorted_data = sorted(raw_data, key=custom_sort)

    csv_data = []
    for row in sorted_data: 

        cash_list = list(range(9))

        cash_list[0] = "2000"
        cash_list[1] = "10"
        cash_list[2] = "01"
        cash_list[3] = "ZCRD"

        zeros_number = 10 - len(str(row[0]))

        cash_list[4] =  ("0" * zeros_number) + str(row[0])
        cash_list[5] =  ("0" * zeros_number) + str(row[0])

        cash_list[6] = "Z01"
        cash_list[7] = f'{row[1]}'
        cash_list[8] = f'{row[2]}'
       
        csv_data.append(cash_list)
    
    csv_data.append(list(range(9)))


    with open (f"{desired_part}.csv", 'w') as csv_obj:  

        writer = csv.writer(csv_obj, delimiter="|", lineterminator="\n")
        # global variables 
        count = 0 
        pointer = 0

        for line in csv_data: 

            if (count > 0) : 
                count -= 1
                continue

            #for more safe data parsing I sliced the line to 9 elements (from index 0 to 8)  to prevent lines issues 
            safe_line = line[:9]

            for next_line in csv_data[pointer:]: 

                if count == 0:
                    count += 1
                    continue

                # if hte next line doesn't match
                if line[4] != next_line[4]:

                    pointer += count
                    count -= 1

                    # case 1 : +0 column added by the program
                    if count == 0:
                        match safe_line[-2]: 

                            case "GCB12": 
                                
                                prefix = ["GCB06", "", "GCB03", ""]

                                safe_line = list_combiner(safe_line, prefix)

                            case "GCB06" : 

                                pfx1 = ["GCB12", ""]
                                pfx2 = ["GCB03", ""]

                                new_slice = list_combiner(safe_line[:7], pfx1, safe_line[-2:], pfx2)

                                safe_line = new_slice

                            case "GCB03": 

                                prefix = ["GCB12", "", "GCB06", ""]

                                new_slice = list_combiner(safe_line[:7], prefix, safe_line[-2:])

                                safe_line = new_slice

                    
                    # case 2 : +2 columns added by the program
                    if count == 1: 

                        match safe_line[-2]:

                            case "GCB06": 
                                prefix = ["GCB03", ""]

                                safe_line = list_combiner(safe_line, prefix)

                            case "GCB03": 

                                # one line condition
                                prefix = ["GCB12", ""] if safe_line[-4] == "GCB06" else ["GCB06", ""]

                                if safe_line[-4] == "GCB06": 
                                    
                                    new_slice = list_combiner(safe_line[:7], prefix, safe_line[-4:])

                                    safe_line = new_slice
                                else : 
                                    new_slice = list_combiner(safe_line[:9], prefix, safe_line[-2:])

                                    safe_line = new_slice

                    # converting 0 to an empty string ""

                    safe_line[7:] = ["" if item == "0" else item for item in safe_line[7:]]

                    # writer line 
                    writer.writerow(safe_line)

                    safe_line = []
                    sliced_next_line = []

                    break


                count += 1

                # append the next_line to the main list (safe_line)

                if line[4] == next_line[4]: 

                    # ['2000', '10', '01', 'ZCRD', '0000100123', '0000100123', 'Z01', 'GCB06', '28']

                    sliced_next_line = list(next_line)[-2:]

                    safe_line = list_combiner(safe_line, sliced_next_line)
                        
                    

        # csv_obj.seek(-1, 2)  # Move the file pointer to one position before the end
        # csv_obj.truncate()   # Truncate the file at the current position



